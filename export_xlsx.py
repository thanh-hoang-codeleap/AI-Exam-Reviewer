import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Color

with open('output.json') as file:
    data = json.load(file)

flattened_data = []

for item in data:
    original = item['original']
    corrected = item['corrected']
    for mistake in item['mistakes']:
        flattened_data.append({
            "Original Sentence": original,
            "Corrected Sentence": corrected,
            "Incorrect Word/ Phrase": mistake['incorrect'],
            "Correction for Word/ Phrase": mistake['correction'],
            "Mistake Short-form": mistake['short_form'],
            "Mistake Analysis": mistake['analysis']
        })

df = pd.DataFrame(flattened_data)

# Save to Excel
df.to_excel("mistakes_analysis.xlsx", index=False)
excel_path = "mistakes_analysis.xlsx"

wb = load_workbook(excel_path)
ws = wb.active

# Function to merge cells in a column if their values are the same
def merge_cells(ws, column):
    prev_value = None
    start_row = None
    
    for row in range(2, ws.max_row + 1):  # Skip header row (row 1)
        cell_value = ws.cell(row=row, column=column).value
        if cell_value == prev_value:
            # If the value is the same as the previous one, keep track of the start row
            if start_row is None:
                start_row = row - 1
        else:
            # If the value has changed, merge the previous cells if needed
            if start_row is not None and start_row < row - 1:
                ws.merge_cells(start_row=start_row, start_column=column, end_row=row - 1, end_column=column)
            # Reset tracking for new value
            start_row = None
        prev_value = cell_value

    # Last merge if needed
    if start_row is not None and start_row < ws.max_row:
        ws.merge_cells(start_row=start_row, start_column=column, end_row=ws.max_row, end_column=column)

# Function to add borders to all cells, with special dashed border between merged rows
def add_borders(ws):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    dashed_border = Border(left=Side(style='dashed'), right=Side(style='dashed'), top=Side(style='dashed'), bottom=Side(style='dashed'))
    
    prev_value = None
    start_row = None

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border  # Apply normal thin border to all cells

        # If the current row has the same value in the first column as the previous one, apply dashed border
        if ws.cell(row=row, column=1).value == prev_value:
            ws.cell(row=row - 1, column=1).border = Border(bottom=Side(style='dashed'))  # Dashed border for the row before this merged group

        # Update the previous value for comparison
        prev_value = ws.cell(row=row, column=1).value

# Function to format the 3rd column based on the 5th column
def format_column_3_based_on_5(ws):
    for row in range(2, ws.max_row + 1):  # Starting from the second row (skip header)
        column_5_value = ws.cell(row=row, column=5).value  # Get value of the 5th column
        
        # Apply color formatting based on the content of the 5th column
        if column_5_value and 'writing' in column_5_value.lower():  # Check if "writing" is in the 5th column
            ws.cell(row=row, column=3).font = Font(color="FF156082")  # Blue font (Hex code for blue)
        else:
            ws.cell(row=row, column=3).font = Font(color="FF0000")  # Red font (Hex code for red)

# Merge the cells in 'original' (column 1) and 'corrected' (column 2)
merge_cells(ws, 1)
merge_cells(ws, 2)

# Add borders to all cells, with dashed border for merged rows
add_borders(ws)

# Apply formatting to the 3rd column based on the 5th column
format_column_3_based_on_5(ws)

# Save the modified Excel file
wb.save(excel_path)
